# utf-8
import datetime
import sqlite3
import openpyxl
from FDataBase import FDataBase
from werkzeug.security import generate_password_hash


# конфигурация
DATABASE = 'pr_ot.db'
DEBUG = True
SECRET_KEY = 'fdgfh78@#5?>gfhf89dx,v06k'


def connect_db():
    conn = sqlite3.connect('pr_ot.db')
    conn.row_factory = sqlite3.Row
    return conn



# получить Excel файл,
# открыть БД, построчно записать таблицу пользователей, изменить пароли в секретные, записать пароли, сохранить БД


def unpacking_file():
    path = "spisok_sotrudnikov.xlsx"  # spisok_sotrudnikov.xlsx sv.xlsx
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj.active
    max_rows = sheet.max_row
    j = 2
    id = 0
    dict_data_user = {}
    for i in range(2, max_rows):
        cell_obj = sheet.cell(row=i, column=j)
        name = cell_obj.value

        cell_obj = sheet.cell(row=i, column=j+1)
        born_date = str(cell_obj.value)[:10]

        cell_obj = sheet.cell(row=i, column=j+2)
        position = cell_obj.value

        cell_obj = sheet.cell(row=i, column=j+3)
        servis = cell_obj.value

        cell_obj = sheet.cell(row=i, column=j+4)
        email = cell_obj.value

        list_data_user = [name, born_date, position, servis, email]
        id += 1
        dict_data_user[id] = list_data_user
    return dict_data_user


# проверка на дубли, если будет полный тезка, то его удалит
def del_double(dict):
    dict_data_user = dict
    data_user = []
    key_lict = []
    for key, value in dict_data_user.items():
        if value[0] not in data_user:
            data_user.append(value[0])
        else:
            key_lict.append(key)
    for key in key_lict:
        del dict_data_user[key]
    return dict_data_user


def del_doubl_in_db(dict_rezult):
    db = connect_db()
    dbase = FDataBase(db)
    dict_db = dbase.all_users()
    if dict_db == "False":
        return dict_rezult
    db_list_user = []
    for user in dict_db:
        full_name = ' '.join(user)
        db_list_user.append(full_name)
    dict_id = dict_rezult
    double_user = []
    for id_user, value in dict_id.items():
        if value[0] in db_list_user:
            double_user.append(id_user)
    for id_user in double_user:
        del dict_id[id_user]
    return dict_id


def save_in_db(dict):
    db = connect_db()
    dbase = FDataBase(db)
    number_id_user = dbase.number_id_user()
    if number_id_user == 'False':
        number_id = 0
    else:
        number_id = number_id_user[0]
    for id_user, value in dict.items():
        full_name = value[0]
        name, firstname, lastname = full_name.split()
        dateborn = value[1]
        position = value[2]
        name_suborganization = value[3]
        email = value[4]
        # Модуль создания пароля
        hpsw = generate_password_hash(dateborn)
        # Запись в БД
        time = datetime.datetime.now()
        role = 1
        id_user = id_user + int(number_id)
        dbase.create_user(id_user, name, firstname, lastname, dateborn, position, name_suborganization, email, hpsw, time,
                          role)


if __name__ == "__main__":
    dict_data_user = unpacking_file()
    dict_data_user = del_double(dict_data_user)
    dict_data_user = del_doubl_in_db(dict_data_user)
    save_in_db(dict_data_user)