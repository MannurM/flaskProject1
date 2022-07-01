# utf-8
import openpyxl
import sqlite3
from FDataBase import FDataBase


# конфигурация
DATABASE = 'pr_ot.db'
DEBUG = True
SECRET_KEY = 'fdgfh78@#5?>gfhf89dx,v06k'


def connect_db():
    conn = sqlite3.connect('pr_ot.db')
    conn.row_factory = sqlite3.Row
    return conn


def unpacking_file():
    path = "static/courses/tests/test_ot.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj.active
    max_rows = sheet.max_row
    j = 2
    id_qestion = 0
    dict_id_qestion = {}
    for i in range(1, max_rows):
        cell_obj = sheet.cell(row=i, column=j)
        qestion = cell_obj.value
        qestion = remove_unwanted(qestion)

        cell_obj_1 = sheet.cell(row=i, column=j + 1)
        answer_1 = cell_obj_1.value
        answer_1 = remove_unwanted(answer_1)
        answer_1 = remove_unwanted_behind(answer_1)

        cell_obj_2 = sheet.cell(row=i, column=j + 2)
        answer_2 = cell_obj_2.value
        answer_2 = remove_unwanted(answer_2)
        answer_2 = remove_unwanted_behind(answer_2)

        cell_obj_3 = sheet.cell(row=i, column=j + 3)
        answer_3 = cell_obj_3.value
        answer_3 = remove_unwanted(answer_3)
        answer_3 = remove_unwanted_behind(answer_3)

        cell_obj_4 = sheet.cell(row=i, column=j + 4)
        answer_4 = cell_obj_4.value
        answer_4 = remove_unwanted(answer_4)
        answer_4 = remove_unwanted_behind(answer_4)

        cell_obj_5 = sheet.cell(row=i, column=j + 5)
        number_answer_just = cell_obj_5.value

        if number_answer_just == 1:
            answer_just = answer_1
        elif number_answer_just == 2:
            answer_just = answer_2
        elif number_answer_just == 3:
            answer_just = answer_3
        elif number_answer_just == 4:
            answer_just = answer_4
        else:
            print('ошибка,  нет правильного ответа', number_answer_just)
            continue
        # Компоновка словаря с данными
        id_qestion += 1
        list_answers = [answer_1, answer_2, answer_3, answer_4]
        dict_id_qestion[id_qestion] = {'qestion': qestion, 'list_answers': list_answers, 'answer_just': answer_just}
    return dict_id_qestion


def remove_unwanted(obj_str):
    list_error_symbol = ['.', ':', ';', ')', '/', ',']  # все значения кроме букв и цифр
    list_error_other = ['a', 'b', 'c', 'd', 'A', 'B', 'C', 'D', '1', '2', '3', '4', 'А', 'Б', 'В', 'Г', 'а',
                        'б', 'в', 'г']  # все буквы и цифры!
    old = None
    answer_new = obj_str

    for i in obj_str:
        if old:
            if i.isalpha() and old.isalpha():
                break
            elif old in list_error_other and i in list_error_symbol:
                answer_new = answer_new.replace(old, '', 1)
                answer_new = answer_new.replace(i, '', 1)
                old = None
            elif i in list_error_symbol:
                answer_new = answer_new.replace(i, '', 1)
            elif old == ' ' and i == ' ':
                answer_new = answer_new.replace(i, '', 2)
            elif old == ' ' and i in list_error_symbol:
                answer_new = answer_new.replace(old, '', 1)
                answer_new = answer_new.replace(i, '', 1)
            elif i.isdigit() and old.isdigit():
                break

        if i == chr(160):
            answer_new = answer_new.replace(i, '', 1)
        else:
            old = i
    return answer_new


def remove_unwanted_behind(obj_str):
    list_error_symbol = ['.', ':', ';', ')', '/', ',']  # все значения кроме букв и цифр
    old = None
    answer_new = obj_str
    for i in reversed(obj_str):
        if old:
            if i.isalpha() and old.isalpha():
                break
        elif i in list_error_symbol:
            answer_new = answer_new.replace(i, '', 1)
        elif i == chr(160):
            answer_new = answer_new.replace(i, '', 1)
        else:
            old = i
    return answer_new


# проверка на дубли
def del_double(dict_id_qestion):
    dict_qestion = dict_id_qestion
    double_key = []
    key_lict = []
    for id_qestion, value in dict_qestion.items():
        for key, val in value.items():
            if key == 'qestion':
                if dict_qestion[id_qestion][key] not in key_lict:
                    key_lict.append(dict_qestion[id_qestion][key])
                else:
                    double_key.append(id_qestion)  # cоздание списка двойных ключей
                    continue
            else:
                continue
    # удаление ключей в словаре из списка дублей
    for id_qestion in double_key:
        del dict_id_qestion[id_qestion]
    return dict_qestion


def del_doubl_in_db(dict_rezult):
    db = connect_db()
    dbase = FDataBase(db)
    dict_db = dbase.all_qestion()
    if dict_db == "False":
        return dict_rezult
    db_list = []
    for r in dict_db:
        db_list.append(r[0])
    dict_id = dict_rezult
    double_qestion = []
    for id_qestion, value in dict_id.items():
        for key, val in value.items():
            if key == 'qestion':
                if val in db_list:
                    print('double')
                    double_qestion.append(id_qestion)
                else:
                    print('No double')
    for id_qestion in double_qestion:
        del dict_id[id_qestion]
    return dict_id


def save_in_db(dict_rezult):
    db = connect_db()
    dbase = FDataBase(db)
    dict_rezult = dict_rezult
    for id_qestion, value in dict_rezult.items():
        qestion_txt, list_answers, answer_just = None, [], None,
        for key, val in value.items():
            if key == 'qestion':
                qestion_txt = dict_rezult[id_qestion][key]
            elif key == 'list_answers':
                list_answers = dict_rezult[id_qestion][key]
            elif key == 'answer_just':
                answer_just = dict_rezult[id_qestion][key]
        dbase.create_test(qestion_txt, list_answers, answer_just)


if __name__ == '__main__':
    dict_id_qestion = unpacking_file()
    dict_rezult = del_double(dict_id_qestion)
    dict_rezult = del_doubl_in_db(dict_rezult)
    save_in_db(dict_rezult)
